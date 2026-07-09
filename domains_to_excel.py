from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Domains"

headers = [
    "User",
    "Main Domain",
    "Sub Domains",
    "Parked Domains",
    "Addon Domains"
]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col)
    c.value = h
    c.font = Font(bold=True)

users = {}

with open("/root/userdatadomains.txt") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue

        domain, rest = line.split(": ", 1)
        parts = rest.split("==")

        user = parts[0]
        dtype = parts[2]
        parent = parts[3]

        if user not in users:
            users[user] = {
                "main": "",
                "sub": [],
                "parked": [],
                "addon": []
            }

        if dtype == "main":
            users[user]["main"] = domain
        elif dtype == "sub":
            users[user]["sub"].append(domain)
        elif dtype == "parked":
            users[user]["parked"].append(domain)
        elif dtype == "addon":
            users[user]["addon"].append(domain)

row = 2

for user in sorted(users):
    ws.cell(row=row, column=1).value = user
    ws.cell(row=row, column=2).value = users[user]["main"]
    ws.cell(row=row, column=3).value = "\n".join(users[user]["sub"])
    ws.cell(row=row, column=4).value = "\n".join(users[user]["parked"])
    ws.cell(row=row, column=5).value = "\n".join(users[user]["addon"])

    for col in range(1, 6):
        ws.cell(row=row, column=col).alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

    row += 1

# Column widths
widths = {
    "A": 20,
    "B": 35,
    "C": 60,
    "D": 40,
    "E": 40,
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width

wb.save("/root/domains_report.xlsx")

print("Created: /root/domains_report.xlsx")
